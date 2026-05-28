"""
数据导出服务 - 生成 Excel 文件
"""
import io

from openpyxl import Workbook
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from backend.models.member import Member, MemberStatus, CardType
from backend.models.order import Order, OrderStatus
from backend.models.booking import Booking, BookingStatus
from backend.models.course import CourseSchedule, Course
from backend.models.lead import Lead


class ExportService:
    CARD_TYPE_MAP = {
        CardType.SINGLE: "次卡",
        CardType.MONTHLY: "月卡",
        CardType.QUARTERLY: "季卡",
        CardType.YEARLY: "年卡",
        CardType.STORED: "储值卡",
    }

    MEMBER_STATUS_MAP = {
        MemberStatus.ACTIVE: "正常",
        MemberStatus.FROZEN: "冻结",
        MemberStatus.SUSPENDED: "休学",
        MemberStatus.CANCELLED: "注销",
    }

    ORDER_STATUS_MAP = {
        OrderStatus.PENDING: "待支付",
        OrderStatus.PAID: "已支付",
        OrderStatus.REFUNDED: "已退款",
        OrderStatus.CANCELLED: "已取消",
    }

    BOOKING_STATUS_MAP = {
        BookingStatus.PENDING: "待确认",
        BookingStatus.CONFIRMED: "已确认",
        BookingStatus.CHECKED_IN: "已签到",
        BookingStatus.CANCELLED: "已取消",
        BookingStatus.NO_SHOW: "未到",
    }

    @staticmethod
    async def export_members(db: AsyncSession, organization_id: int) -> bytes:
        query = select(Member).where(Member.organization_id == organization_id).order_by(Member.created_at.desc())
        result = await db.execute(query)
        members = result.scalars().all()

        wb = Workbook()
        ws = wb.active
        ws.title = "会员数据"

        headers = ["ID", "姓名", "手机号", "邮箱", "性别", "卡种", "状态", "剩余次数", "储值余额",
                    "开卡日期", "到期日期", "累计消费", "等级", "创建时间"]
        ws.append(headers)

        for m in members:
            ws.append([
                m.id, m.name, m.phone, m.email or "",
                m.gender or "",
                ExportService.CARD_TYPE_MAP.get(m.card_type, m.card_type.value if m.card_type else ""),
                ExportService.MEMBER_STATUS_MAP.get(m.status, m.status.value if m.status else ""),
                m.card_remaining_count or 0,
                m.card_balance or 0,
                m.card_start_date.strftime("%Y-%m-%d") if m.card_start_date else "",
                m.card_end_date.strftime("%Y-%m-%d") if m.card_end_date else "",
                m.total_consumption or 0,
                m.level or 1,
                m.created_at.strftime("%Y-%m-%d %H:%M") if m.created_at else "",
            ])

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    @staticmethod
    async def export_orders(db: AsyncSession, organization_id: int) -> bytes:
        query = select(Order).where(Order.organization_id == organization_id).order_by(Order.created_at.desc())
        result = await db.execute(query)
        orders = result.scalars().all()

        wb = Workbook()
        ws = wb.active
        ws.title = "订单数据"

        headers = ["ID", "订单号", "会员ID", "金额", "实付金额", "支付方式",
                    "支付状态", "创建时间", "支付时间"]
        ws.append(headers)

        for o in orders:
            ws.append([
                o.id, o.order_no, o.member_id,
                o.amount, o.actual_amount,
                o.payment_method or "",
                ExportService.ORDER_STATUS_MAP.get(o.payment_status, o.payment_status.value if o.payment_status else ""),
                o.created_at.strftime("%Y-%m-%d %H:%M") if o.created_at else "",
                o.paid_at.strftime("%Y-%m-%d %H:%M") if o.paid_at else "",
            ])

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    @staticmethod
    async def export_bookings(db: AsyncSession, organization_id: int) -> bytes:
        query = (
            select(Booking, CourseSchedule, Course)
            .join(CourseSchedule, Booking.schedule_id == CourseSchedule.id)
            .join(Course, CourseSchedule.course_id == Course.id)
            .where(Booking.organization_id == organization_id)
            .order_by(Booking.created_at.desc())
        )
        result = await db.execute(query)
        rows = result.all()

        wb = Workbook()
        ws = wb.active
        ws.title = "预约数据"

        headers = ["ID", "会员ID", "课程名称", "上课时间", "预约状态", "签到时间", "签到方式", "创建时间"]
        ws.append(headers)

        for booking, schedule, course in rows:
            ws.append([
                booking.id, booking.member_id, course.name,
                schedule.start_time.strftime("%Y-%m-%d %H:%M") if schedule.start_time else "",
                ExportService.BOOKING_STATUS_MAP.get(booking.status, booking.status.value if booking.status else ""),
                booking.check_in_time.strftime("%Y-%m-%d %H:%M") if booking.check_in_time else "",
                booking.check_in_method or "",
                booking.created_at.strftime("%Y-%m-%d %H:%M") if booking.created_at else "",
            ])

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    @staticmethod
    async def export_leads(db: AsyncSession, organization_id: int) -> bytes:
        query = select(Lead).where(Lead.organization_id == organization_id).order_by(Lead.created_at.desc())
        result = await db.execute(query)
        leads = result.scalars().all()

        wb = Workbook()
        ws = wb.active
        ws.title = "潜客数据"

        headers = ["ID", "姓名", "手机号", "来源", "状态", "意向", "预算",
                    "跟进次数", "最后联系", "创建时间"]
        ws.append(headers)

        for lead in leads:
            ws.append([
                lead.id, lead.name, lead.phone or "",
                lead.source or "", lead.status or "",
                lead.intent or "", lead.expected_budget or "",
                lead.follow_up_count, lead.last_contacted_at.strftime("%Y-%m-%d %H:%M") if lead.last_contacted_at else "",
                lead.created_at.strftime("%Y-%m-%d %H:%M") if lead.created_at else "",
            ])

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()
